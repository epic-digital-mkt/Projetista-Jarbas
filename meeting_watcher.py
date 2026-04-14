"""
meeting_watcher.py
==================
Monitora o ClickUp em background. Quando um novo doc de reunião é detectado,
gera Use Cases automaticamente: Gemini -> Excel -> Tarefa no ClickUp com anexo.

Agendado via Windows Task Scheduler (a cada 1 hora).
"""

import os, json, re, time, urllib.request, urllib.error, urllib.parse
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─── Carrega .env ─────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
_env = BASE_DIR / ".env"
if _env.exists():
    for _line in _env.read_text(encoding="utf-8").splitlines():
        if "=" in _line and not _line.startswith("#"):
            _k, _v = _line.split("=", 1)
            os.environ.setdefault(_k.strip(), _v.strip())

# ─── Configurações ────────────────────────────────────────────────────────────
CLICKUP_API_KEY   = os.environ["CLICKUP_API_KEY"]
GEMINI_API_KEY    = os.environ.get("GEMINI_API_KEY", "")
GEMINI_MODELS     = ["gemini-2.5-flash", "gemini-2.5-flash-lite", "gemini-3-flash-preview"]
CLICKUP_WORKSPACE = os.environ["CLICKUP_WORKSPACE"]
CLICKUP_USER_ID   = int(os.environ.get("CLICKUP_USER_ID", "0"))
OUTPUT_DIR        = Path(os.environ.get("OUTPUT_DIR", str(Path.home() / "Downloads" / "Use Cases")))
PROCESSED_FILE    = BASE_DIR / "processed_meetings.json"
LOG_FILE          = BASE_DIR / "watcher.log"

# IDs dos Spaces monitorados: {"space_id": "Nome do Cliente"}
# Preencha com os IDs reais do seu workspace ClickUp
WATCHED_SPACES = {
    # "SPACE_ID_1": "Cliente A",
    # "SPACE_ID_2": "Cliente B",
    # "SPACE_ID_3": "Interno",
}

# Mapeamento Space -> Lista de tarefas: {"space_id": "list_id"}
# Se vazio, o watcher busca automaticamente uma lista com "reunião" no nome
TASK_LISTS = {
    # "SPACE_ID_1": "LIST_ID_1",
}

MEETING_KEYWORDS = [
    "weekly", "discovery", "pds", "alinhamento",
    "pedidos", "novo", "proposta", "projeto"
]

# ─── Constantes ───────────────────────────────────────────────────────────────
CLICKUP_PARENT_TYPE_SPACE = 4
STATUS_OPTIONS = ["A fazer", "Em andamento", "Concluído", "Backlog", "Impedido"]
MOSCOW_OPTIONS = ["Must", "Should", "Could", "Won't"]
SHIRT_SIZES    = ["S", "M", "L", "XL"]
GEMINI_CONTENT_LIMIT = 15000

# ─── Regex pré-compilados ─────────────────────────────────────────────────────
_RE_UNSAFE_CHARS = re.compile(r'[<>:"/\\|?*]')
_RE_DATE_DMY     = re.compile(r'(\d{2})[/-](\d{2})[/-](\d{4})')
_RE_DATE_YMD     = re.compile(r'(\d{4})[/-](\d{2})[/-](\d{2})')
_RE_JSON_FENCE   = re.compile(r'```json\s*|\s*```')

# ─── Log ──────────────────────────────────────────────────────────────────────

def log(msg):
    line = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
    print(line)
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass

# ─── Controle de processados ──────────────────────────────────────────────────

def load_processed():
    if PROCESSED_FILE.exists():
        try:
            return json.loads(PROCESSED_FILE.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            log("AVISO: processed_meetings.json corrompido — reiniciando estado.")
    return {}

def save_processed(data):
    PROCESSED_FILE.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

# ─── HTTP com retry ───────────────────────────────────────────────────────────

def _do_request(req, timeout=30):
    """Requisição HTTP com retry automático para 429/503."""
    for attempt in range(4):
        try:
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return json.loads(r.read())
        except urllib.error.HTTPError as e:
            if e.code in (429, 503) and attempt < 3:
                wait = 2 ** (attempt + 2)
                log(f"  Serviço ocupado ({e.code}) — aguardando {wait}s...")
                time.sleep(wait)
                continue
            raise
    raise RuntimeError("Máximo de tentativas atingido")

# ─── ClickUp helpers ──────────────────────────────────────────────────────────

def cu_get(path):
    req = urllib.request.Request(
        f"https://api.clickup.com{path}",
        headers={"Authorization": CLICKUP_API_KEY}
    )
    return _do_request(req, timeout=30)

def cu_post(path, data):
    payload = json.dumps(data).encode("utf-8")
    req = urllib.request.Request(
        f"https://api.clickup.com{path}",
        data=payload, method="POST",
        headers={"Authorization": CLICKUP_API_KEY, "Content-Type": "application/json"}
    )
    return _do_request(req, timeout=30)

def cu_attach(task_id, file_path, filename):
    file_data = Path(file_path).read_bytes()
    filename_safe = urllib.parse.quote(filename, safe="._- ")
    boundary = b"----WatcherBoundary99887766"
    body = (
        b"--" + boundary + b"\r\n"
        + f'Content-Disposition: form-data; name="attachment"; filename="{filename_safe}"\r\n'.encode()
        + b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
        + file_data + b"\r\n--" + boundary + b"--\r\n"
    )
    req = urllib.request.Request(
        f"https://api.clickup.com/api/v2/task/{task_id}/attachment",
        data=body, method="POST",
        headers={"Authorization": CLICKUP_API_KEY,
                 "Content-Type": f"multipart/form-data; boundary={boundary.decode()}"}
    )
    return _do_request(req, timeout=60)

# ─── Gemini helpers compartilhados ────────────────────────────────────────────

def _gemini_request(prompt, temperature=0.2, max_tokens=16384, timeout=90):
    """Chama a API Gemini com fallback automático entre modelos."""
    if not GEMINI_API_KEY:
        return None
    payload = json.dumps({
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": temperature, "maxOutputTokens": max_tokens}
    }).encode("utf-8")
    last_err = None
    for model in GEMINI_MODELS:
        req = urllib.request.Request(
            f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={GEMINI_API_KEY}",
            data=payload, method="POST",
            headers={"Content-Type": "application/json"}
        )
        try:
            return _do_request(req, timeout=timeout)
        except urllib.error.HTTPError as e:
            last_err = e
            if e.code in (503, 429):
                log(f"  Modelo {model} indisponível ({e.code}), tentando próximo...")
                continue
            raise
    raise last_err

def _parse_gemini_json(result):
    """Extrai e repara JSON do resultado da API Gemini."""
    try:
        text = result["candidates"][0]["content"]["parts"][0]["text"]
    except (KeyError, IndexError, TypeError):
        return None
    text = _RE_JSON_FENCE.sub("", text).strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        last_brace = text.rfind("}")
        if last_brace != -1:
            try:
                return json.loads(text[:last_brace + 1] + "\n  ]\n}")
            except json.JSONDecodeError:
                pass
    return None

# ─── Busca docs novos ─────────────────────────────────────────────────────────

def is_meeting(doc_name):
    name_lower = doc_name.lower()
    return any(k in name_lower for k in MEETING_KEYWORDS)

def get_new_meeting_docs(processed):
    new_docs, cursor = [], None
    while True:
        url = f"/api/v3/workspaces/{CLICKUP_WORKSPACE}/docs?limit=100"
        if cursor:
            url += f"&next_cursor={urllib.parse.quote(str(cursor))}"
        try:
            data = cu_get(url)
        except Exception as e:
            log(f"Erro ao listar docs: {e}")
            break
        docs = data.get("docs", [])
        for doc in docs:
            doc_id = doc.get("id", "")
            if doc_id not in processed and is_meeting(doc.get("name", "")):
                new_docs.append(doc)
        cursor = data.get("next_cursor")
        if not cursor or not docs:
            break
    return new_docs

# ─── Gemini — gera casos de uso ───────────────────────────────────────────────

def call_gemini(content, meeting_name):
    prompt = f"""Com base no conteúdo abaixo, crie casos de uso no formato JSON.

Retorne APENAS JSON válido (sem markdown), estrutura:
{{
  "casos_de_uso": [
    {{
      "id": "UC01",
      "grupo": "categoria do caso",
      "caso_de_uso": "nome curto",
      "user_story": "Como X, quero Y, para Z.",
      "criterios_aceite": "- Dado X, quando Y, então Z.\\n- ...",
      "detalhes": "contexto e solução técnica",
      "status": "A fazer",
      "responsavel": "nome",
      "shirt_size": "M",
      "horas": "8h",
      "moscow": "Must",
      "inicio": "DD/MM",
      "fim": "DD/MM"
    }}
  ]
}}

Regras:
- shirt_size: {", ".join(SHIRT_SIZES)}
- moscow: {", ".join(MOSCOW_OPTIONS)}
- status: {", ".join(STATUS_OPTIONS)}
- horas: número seguido de h (ex: 4h, 8h, 16h)
- criterios_aceite: lista de bullet points no formato Dado/Quando/Então

--- CONTEÚDO: {meeting_name} ---
{content[:GEMINI_CONTENT_LIMIT]}
"""
    try:
        result = _gemini_request(prompt, temperature=0.2, max_tokens=16384, timeout=90)
        return _parse_gemini_json(result)
    except Exception as e:
        log(f"  Gemini erro: {e}")
        return None

# ─── Recomendação técnica HubSpot ─────────────────────────────────────────────

def ask_david(uc):
    """Lê os dados do UC e retorna dict com sugestao e hub_licenca."""
    prompt = f"""Você é um especialista em HubSpot com conhecimento de documentação oficial e implementações reais.

Com base no caso de uso abaixo, retorne APENAS JSON válido (sem markdown):
{{
  "sugestao": "qual ferramenta/módulo HubSpot usar e quais equipes devem se envolver (máx 3 linhas, texto corrido)",
  "hub_licenca": "Hub e nível de licença necessários. Ex: Sales Hub Pro + Ops Hub Starter"
}}

--- CASO DE USO ---
ID: {uc.get('id','')}
Caso: {uc.get('caso_de_uso','')}
User Story: {uc.get('user_story','')}
Detalhes: {uc.get('detalhes','')}
Critérios de Aceite: {uc.get('criterios_aceite','')}
"""
    try:
        result = _gemini_request(prompt, temperature=0.1, max_tokens=2048, timeout=30)
        data = _parse_gemini_json(result)
        if data:
            return {"sugestao": data.get("sugestao","—"), "hub_licenca": data.get("hub_licenca","—")}
    except Exception as e:
        log(f"    Erro UC {uc.get('id','?')}: {e}")
    return {"sugestao": "— Revisar manualmente.", "hub_licenca": "—"}

# ─── Excel ────────────────────────────────────────────────────────────────────

_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

def create_excel(meeting_name, use_cases_data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Casos de Uso"

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    fills = {
        "title":        PatternFill("solid", fgColor="1F4E79"),
        "header":       PatternFill("solid", fgColor="2E75B6"),
        "david":        PatternFill("solid", fgColor="FFF2CC"),
        "licenca":      PatternFill("solid", fgColor="E2EFDA"),
        "alt":          PatternFill("solid", fgColor="D6E4F0"),
        "Must":         PatternFill("solid", fgColor="C6EFCE"),
        "Should":       PatternFill("solid", fgColor="FFEB9C"),
        "Could":        PatternFill("solid", fgColor="FFD7A8"),
        "Won't":        PatternFill("solid", fgColor="D9D9D9"),
        "Concluído":    PatternFill("solid", fgColor="C6EFCE"),
        "Em andamento": PatternFill("solid", fgColor="BDD7EE"),
        "Atrasado":     PatternFill("solid", fgColor="FFC7CE"),
        "Impedido":     PatternFill("solid", fgColor="FFC7CE"),
        "Backlog":      PatternFill("solid", fgColor="EEEEEE"),
    }

    # Título
    for row, val, height, font_kwargs, fill_key in [
        (1, f"Casos de Uso — {meeting_name}", 30,
         {"bold": True, "color": "FFFFFF", "size": 13}, "title"),
        (2, f"Gerado automaticamente em {datetime.now().strftime('%d/%m/%Y %H:%M')}", 16,
         {"italic": True, "color": "1F4E79", "size": 10}, None),
    ]:
        ws.merge_cells(f"A{row}:T{row}")
        c = ws[f"A{row}"]
        c.value = val
        c.font = Font(name="Calibri", **font_kwargs)
        c.fill = fills[fill_key] if fill_key else PatternFill("solid", fgColor="EBF3FB")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = height

    # Cabeçalho — 20 colunas (A–T)
    headers = ["ID", "Grupo", "Caso de Uso", "User Story", "Detalhes / Solução",
               "Critérios de Aceite", "Sugestão Técnica", "Hub / Licença",
               "Status", "Responsável", "Shirt", "Horas", "MoSCoW", "Início", "Fim",
               "Lista ou ID", "Título", "Descrição (contexto)", "Responsável (Assignee)", "Status (tarefa)"]
    special = {"Sugestão Técnica": ("FFF2CC", "1F4E79"), "Hub / Licença": ("E2EFDA", "1F4E79")}
    task_cols = {"Lista ou ID", "Título", "Descrição (contexto)", "Responsável (Assignee)", "Status (tarefa)"}
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        if h in special:
            bg, fg = special[h]
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(name="Calibri", bold=True, color=fg, size=10)
        elif h in task_cols:
            cell.fill = PatternFill("solid", fgColor="44546A")
            cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        else:
            cell.fill = fills["header"]
            cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[3].height = 22

    # Dropdowns de validação
    dv_status = DataValidation(type="list", formula1=f'"{",".join(STATUS_OPTIONS)}"',  showDropDown=False)
    dv_shirt  = DataValidation(type="list", formula1=f'"{",".join(SHIRT_SIZES)}"',     showDropDown=False)
    dv_moscow = DataValidation(type="list", formula1=f'"{",".join(MOSCOW_OPTIONS)}"',  showDropDown=False)
    for dv in [dv_status, dv_shirt, dv_moscow]:
        ws.add_data_validation(dv)

    # Linhas de dados
    casos = use_cases_data.get("casos_de_uso", []) if use_cases_data else []
    for ri, uc in enumerate(casos, 4):
        base_fill = fills["alt"] if ri % 2 == 0 else _WHITE_FILL
        vals = [
            uc.get("id",""),             uc.get("grupo",""),          uc.get("caso_de_uso",""),
            uc.get("user_story",""),     uc.get("detalhes",""),       uc.get("criterios_aceite",""),
            uc.get("sugestao_david",""), uc.get("hub_licenca",""),
            uc.get("status",""),         uc.get("responsavel",""),    uc.get("shirt_size",""),
            uc.get("horas",""),          uc.get("moscow",""),         uc.get("inicio",""), uc.get("fim",""),
            "", "", "", "", ""
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font   = Font(name="Calibri", size=10, bold=(ci in [9, 13, 20]))
            cell.border = thin
            cell.alignment = Alignment(
                horizontal="center" if ci in [1,9,10,11,12,13,14,15,16,19,20] else "left",
                vertical="center", wrap_text=True
            )
            if   ci == 7:  cell.fill = fills["david"]
            elif ci == 8:  cell.fill = fills["licenca"]
            elif ci == 13: cell.fill = fills.get(val, base_fill)
            elif ci == 9:  cell.fill = fills.get(val, base_fill)
            elif ci >= 16: cell.fill = _WHITE_FILL
            else:          cell.fill = base_fill

        dv_status.add(ws.cell(row=ri, column=9))
        dv_shirt.add(ws.cell(row=ri,  column=11))
        dv_moscow.add(ws.cell(row=ri, column=13))
        ws.row_dimensions[ri].height = 60

    for i, w in enumerate([8,16,28,42,46,40,50,28,14,18,9,8,10,9,9,16,28,40,18,14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

# ─── Extrai data da reunião ───────────────────────────────────────────────────

def extract_due_date_ms(meeting_name, fallback_ts_ms=None):
    m = _RE_DATE_YMD.search(meeting_name)
    if m:
        try:
            dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            return int(dt.timestamp() * 1000)
        except (ValueError, OverflowError):
            pass
    m = _RE_DATE_DMY.search(meeting_name)
    if m:
        try:
            a, b, c = int(m.group(1)), int(m.group(2)), int(m.group(3))
            dt = datetime(c, b, a) if a > 12 else datetime(c, a, b)
            return int(dt.timestamp() * 1000)
        except (ValueError, OverflowError):
            pass
    return fallback_ts_ms

# ─── Detecta lista de tarefas ─────────────────────────────────────────────────

def find_task_list(space_id):
    if space_id in TASK_LISTS:
        return TASK_LISTS[space_id]
    try:
        data = cu_get(f"/api/v2/space/{space_id}/list?archived=false")
        for lst in data.get("lists", []):
            if any(x in lst["name"].lower() for x in ["reunião", "reuniao", "meeting"]):
                TASK_LISTS[space_id] = lst["id"]
                return lst["id"]
    except Exception as e:
        log(f"  Erro ao buscar listas do espaço {space_id}: {e}")
    log(f"  AVISO: espaço {space_id} não mapeado e nenhuma lista encontrada.")
    return None

# ─── Processa uma reunião ─────────────────────────────────────────────────────

def _infer_space_from_name(name):
    """Infere o space_id com base no nome do doc. Personalize conforme seus clientes."""
    name_lower = name.lower()
    for space_id, space_name in WATCHED_SPACES.items():
        if space_name.lower() in name_lower:
            return space_id
    return None

def process_meeting(doc):
    doc_id = doc["id"]
    name   = doc.get("name", "Reunião")
    parent = doc.get("parent", {})

    parent_type = parent.get("type") if isinstance(parent, dict) else None
    raw_id      = parent.get("id")   if isinstance(parent, dict) else None
    space_id = (raw_id if parent_type == CLICKUP_PARENT_TYPE_SPACE and raw_id in WATCHED_SPACES
                else _infer_space_from_name(name))

    log(f"Processando: {name}")

    try:
        pages_resp = cu_get(f"/api/v3/workspaces/{CLICKUP_WORKSPACE}/docs/{doc_id}/pages")
        if isinstance(pages_resp, dict):
            pages = pages_resp.get("pages", [])
        elif isinstance(pages_resp, list):
            pages = pages_resp
        else:
            pages = []
        content = pages[0].get("content", "") if pages else ""
    except Exception as e:
        log(f"  Erro ao ler doc: {e}")
        return None

    if not content.strip():
        log("  Doc vazio — ignorando.")
        return None

    use_cases = call_gemini(content, name)
    if use_cases:
        log(f"  Gemini: {len(use_cases.get('casos_de_uso', []))} casos de uso gerados.")
    else:
        log("  Gemini indisponível — estrutura básica.")
        use_cases = {"casos_de_uso": [{
            "id": "UC01", "grupo": "Geral",
            "caso_de_uso": "Revisar manualmente",
            "user_story": f"Transcrição disponível no ClickUp: {name}",
            "criterios_aceite": "— Revisar manualmente.",
            "detalhes": "Gemini não disponível no momento.",
            "sugestao_david": "— Revisar manualmente.",
            "hub_licenca": "—",
            "status": STATUS_OPTIONS[0], "responsavel": "—",
            "shirt_size": "M", "horas": "—", "moscow": MOSCOW_OPTIONS[0],
            "inicio": datetime.now().strftime("%d/%m"), "fim": "—"
        }]}

    # Recomendação técnica — chamadas paralelas, uma por UC
    if GEMINI_API_KEY:
        casos = use_cases.get("casos_de_uso", [])
        with ThreadPoolExecutor(max_workers=min(len(casos), 5)) as pool:
            futures = {pool.submit(ask_david, uc): uc for uc in casos}
            for future in as_completed(futures):
                uc = futures[future]
                david = future.result()
                uc["sugestao_david"] = david.get("sugestao", "—")
                uc["hub_licenca"]    = david.get("hub_licenca", "—")
                log(f"  Recomendação -> {uc.get('id','?')}: ok")

    safe_name = _RE_UNSAFE_CHARS.sub('_', name)
    xlsx_name = f"{safe_name}_Use_Cases.xlsx"
    xlsx_path = OUTPUT_DIR / xlsx_name
    create_excel(name, use_cases, xlsx_path)
    log(f"  Excel: {xlsx_path}")

    list_id = find_task_list(space_id)
    if not list_id:
        log("  ERRO: nenhuma lista de tarefas encontrada. Configure TASK_LISTS no .env.")
        return None

    desc_lines = ["Casos de uso gerados automaticamente.\n"] + [
        f"**{uc['id']} — {uc['caso_de_uso']}**\n"
        f"  {uc.get('responsavel','—')} | {uc.get('horas','—')} | {uc.get('moscow','—')} | até {uc.get('fim','—')}"
        for uc in use_cases.get("casos_de_uso", [])
    ]

    created_ms  = int(doc.get("date_created", 0))
    due_date_ms = extract_due_date_ms(name, fallback_ts_ms=created_ms)
    task_payload = {
        "name": f"{name} - Use Cases",
        "description": "\n\n".join(desc_lines),
        "status": "to do",
        "priority": 2,
        "notify_all": False,
    }
    if CLICKUP_USER_ID:
        task_payload["assignees"] = [CLICKUP_USER_ID]
    if due_date_ms:
        task_payload["due_date"] = due_date_ms
        task_payload["due_date_time"] = False

    try:
        task     = cu_post(f"/api/v2/list/{list_id}/task", task_payload)
        task_id  = task.get("id")
        task_url = task.get("url")
        log(f"  Tarefa: {task_url}")
    except Exception as e:
        log(f"  Erro ao criar tarefa: {e}")
        return None

    try:
        cu_attach(task_id, xlsx_path, xlsx_name)
        log("  Excel anexado.")
    except Exception as e:
        log(f"  AVISO: falha ao anexar Excel — {e}")

    return task_url

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    log("=== Watcher iniciado ===")
    processed = load_processed()
    new_docs  = get_new_meeting_docs(processed)

    if not new_docs:
        log("Nenhuma reunião nova encontrada.")
        return

    log(f"{len(new_docs)} reunião(ões) nova(s) encontrada(s).")
    for doc in new_docs:
        try:
            task_url = process_meeting(doc)
            if task_url:
                processed[doc["id"]] = {
                    "name": doc.get("name", ""),
                    "task_url": task_url,
                    "processed_at": datetime.now().isoformat()
                }
        except Exception as e:
            log(f"  ERRO ao processar '{doc.get('name','')}': {e}")

    save_processed(processed)
    log("=== Watcher concluído ===\n")

if __name__ == "__main__":
    main()
